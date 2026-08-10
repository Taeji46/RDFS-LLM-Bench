"""Aggregate per-entry eval results into flat reports for analysis.

Reads eval JSONL files from:
  data/llm-eval/eval/{response_type}/{model}/{op}/{dataset}/{n-rule}/
    eval__{model}__{op}__{dataset}__{rule}__n{N}__...__{run_id}.jsonl

Aggregates across ALL response types (openai-batch, sequential, etc.)
and writes:
  data/llm-eval/reports/{mode}/csv/scores-{mode}.csv   (canonical, machine-readable)
  data/llm-eval/reports/{mode}/xlsx/scores-{mode}.xlsx  (human-readable view)

Columns:
  model, prompting_condition, dataset_variant, pattern_id, n_rule,
  n_total, n_correct, accuracy,
  precision_triple, recall_triple, f1_triple,
  [precision_rule, recall_rule, f1_rule]  -- only for ARP-full/ARP-name
"""

from __future__ import annotations

import argparse
import csv
import sys
from fractions import Fraction
from pathlib import Path

THIS_FILE = Path(__file__).resolve()
LLM_EVAL_DIR = THIS_FILE.parents[1]
PROJECT_ROOT = THIS_FILE.parents[3]
sys.path.insert(0, str(LLM_EVAL_DIR))

from shared.io import read_jsonl
from shared.numeric import mean_fraction, metric_from_str, metric_to_excel_float, metric_to_str

DEFAULT_EVAL_ROOT    = PROJECT_ROOT / "data" / "llm-eval" / "eval"
DEFAULT_REPORT_ROOT  = PROJECT_ROOT / "data" / "llm-eval" / "reports"


class EvalValidationError(ValueError):
    """Validation error that should stop aggregation for the whole run."""

    def __init__(self, issue: str, message: str) -> None:
        super().__init__(message)
        self.issue = issue


def _relpath(path: Path) -> str:
    try:
        return path.relative_to(PROJECT_ROOT).as_posix()
    except ValueError:
        return str(path)


def _csv_items(text: str) -> list[str]:
    return [s.strip() for s in text.split(",") if s.strip()]


def _expand_rules(pattern_id: str) -> list[str]:
    """Expand composite pattern_id into constituent rules.

    Examples:
      "rdfs2"    -> ["rdfs2"]
      "rdfs2_3"  -> ["rdfs2", "rdfs3"]
      "rdfs9_11" -> ["rdfs9", "rdfs11"]
    """
    import re
    m = re.match(r'^(rdfs)(\d+)((?:_\d+)*)$', pattern_id)
    if not m:
        return [pattern_id]
    prefix, first, rest = m.group(1), m.group(2), m.group(3)
    rules = [f"{prefix}{first}"]
    for num in rest.lstrip("_").split("_"):
        if num:
            rules.append(f"{prefix}{num}")
    return rules


def _parse_eval_filename(stem: str) -> dict | None:
    """Parse eval-{mode}__{model}__{op}__{dataset}__{rule}__n{N}__... stem.

    Returns dict with model, prompting_condition, dataset_variant, pattern_id, rules, n_rule,
    or None if unparseable.
    """
    import re as _re
    m = _re.match(r'^eval(?:-\w+)?__', stem)
    if not m:
        return None
    body = stem[m.end():]
    fields = body.split("__")
    if len(fields) < 4:
        return None
    model          = fields[0]
    prompting_condition = fields[1]
    dataset_variant   = fields[2]
    pattern_id        = fields[3]
    rules          = _expand_rules(pattern_id)
    n_rule         = len(rules)
    op_parts       = prompting_condition.split("-", 1)
    presented_rule_type      = op_parts[0] if len(op_parts) == 2 else prompting_condition
    rule_format      = op_parts[1] if len(op_parts) == 2 else ""
    return {
        "model":            model,
        "prompting_condition":   prompting_condition,
        "presented_rule_type": presented_rule_type,
        "rule_format":        rule_format,
        "dataset_variant":     dataset_variant,
        "pattern_id":          pattern_id,
        "rules":            ",".join(rules),
        "n_rule":           n_rule,
    }


def aggregate_file(eval_path: Path) -> dict | None:
    """Aggregate one eval JSONL into a single summary row."""
    info = _parse_eval_filename(eval_path.stem)
    if info is None:
        return None

    rows = read_jsonl(eval_path)
    if not rows:
        return None

    missing_target_empty_count = sum(1 for r in rows if "target_empty" not in r)
    if missing_target_empty_count:
        raise EvalValidationError(
            "missing_target_empty",
            f"{_relpath(eval_path)} has {missing_target_empty_count}/{len(rows)} rows "
            "without target_empty; re-run evaluate_outputs.py with --overwrite",
        )

    target_empty_count = sum(1 for r in rows if r.get("target_empty"))
    if target_empty_count:
        raise EvalValidationError(
            "target_empty_true",
            f"{_relpath(eval_path)} has {target_empty_count}/{len(rows)} rows "
            "with target_empty=true; inspect the corresponding data/task entries",
        )

    n_total   = len(rows)
    n_correct = sum(1 for r in rows if r.get("overall_ok"))

    def avg(key: str) -> str:
        value = mean_fraction(
            metric
            for r in rows
            if (metric := metric_from_str(r.get(key))) is not None
        )
        if value is None:
            raise ValueError(f"{_relpath(eval_path)} has no values for {key}")
        return metric_to_str(value)

    record: dict = {
        **info,
        "n_total":          n_total,
        "n_correct":        n_correct,
        "accuracy":         metric_to_str(Fraction(n_correct, n_total)) if n_total else "0",
        "precision_triple": avg("precision_triple"),
        "recall_triple":    avg("recall_triple"),
        "f1_triple":        avg("f1_triple"),
    }

    # Rule metrics present only for ARP-full / ARP-name
    if "precision_rule" in rows[0]:
        record["precision_rule"] = avg("precision_rule")
        record["recall_rule"]    = avg("recall_rule")
        record["f1_rule"]        = avg("f1_rule")
    else:
        record["precision_rule"] = ""
        record["recall_rule"]    = ""
        record["f1_rule"]        = ""

    return record


def _print_eval_validation_errors(errors: dict[str, list[str]]) -> None:
    missing = errors.get("missing_target_empty", [])
    if missing:
        print(
            f"[ERROR] {len(missing)} eval files are missing target_empty. "
            "They were produced by the old evaluator; re-run evaluate_outputs.py with --overwrite."
        )
        for message in missing[:10]:
            print(f"  - {message}")
        if len(missing) > 10:
            print(f"  ... {len(missing) - 10} more files")

    target_empty = errors.get("target_empty_true", [])
    if target_empty:
        print(
            f"[ERROR] {len(target_empty)} eval files contain target_empty=true. "
            "These indicate empty scoring targets and must be inspected before aggregation."
        )
        for message in target_empty[:10]:
            print(f"  - {message}")
        if len(target_empty) > 10:
            print(f"  ... {len(target_empty) - 10} more files")

    known = {"missing_target_empty", "target_empty_true"}
    for issue, messages in sorted(errors.items()):
        if issue in known:
            continue
        print(f"[ERROR] {len(messages)} eval files failed validation with issue={issue}.")
        for message in messages[:10]:
            print(f"  - {message}")
        if len(messages) > 10:
            print(f"  ... {len(messages) - 10} more files")


FIELDNAMES = [
    "model", "prompting_condition", "presented_rule_type", "rule_format",
    "dataset_variant", "pattern_id", "rules", "n_rule",
    "n_total", "n_correct", "accuracy",
    "precision_triple", "recall_triple", "f1_triple",
    "precision_rule", "recall_rule", "f1_rule",
]

NUMERIC_FIELDS = {
    "n_rule", "n_total", "n_correct", "accuracy",
    "precision_triple", "recall_triple", "f1_triple",
    "precision_rule", "recall_rule", "f1_rule",
}


def _coerce_excel_value(field: str, value):
    if value == "":
        return ""
    if field in {"n_rule", "n_total", "n_correct"}:
        return int(value)
    if field in NUMERIC_FIELDS:
        return metric_to_excel_float(value)
    return value


def write_excel_view(records: list[dict], out_path: Path) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl is not installed; skipped Excel view.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "scores"

    hdr_fill = PatternFill("solid", fgColor="4472C4")
    hdr_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center")
    right = Alignment(horizontal="right")

    for col, field in enumerate(FIELDNAMES, 1):
        cell = ws.cell(row=1, column=col, value=field)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center

    for row_idx, record in enumerate(records, 2):
        for col, field in enumerate(FIELDNAMES, 1):
            value = _coerce_excel_value(field, record.get(field, ""))
            cell = ws.cell(row=row_idx, column=col, value=value)
            if field in {"accuracy", "precision_triple", "recall_triple", "f1_triple",
                         "precision_rule", "recall_rule", "f1_rule"} and value != "":
                cell.number_format = "0.000"
                cell.alignment = right
            elif field in {"n_rule", "n_total", "n_correct"}:
                cell.number_format = "0"
                cell.alignment = right

    widths = {
        "model": 24,
        "prompting_condition": 20,
        "presented_rule_type": 18,
        "rule_format": 14,
        "dataset_variant": 16,
        "pattern_id": 16,
        "rules": 22,
    }
    for col, field in enumerate(FIELDNAMES, 1):
        ws.column_dimensions[get_column_letter(col)].width = widths.get(field, 13)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Written Excel view -> {_relpath(out_path)}")


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Aggregate eval JSONL files into a flat CSV."
    )
    parser.add_argument("--mode", type=str, default="strict",
                        choices=["strict", "flex"],
                        help="Evaluation mode (default: strict)")
    parser.add_argument("--eval-root", type=Path, default=DEFAULT_EVAL_ROOT)
    parser.add_argument("--report-root", type=Path, default=DEFAULT_REPORT_ROOT)
    parser.add_argument("--response-types", type=str, default="",
                        help="Comma-separated response types to include (default: all)")
    parser.add_argument("--models", type=str, default="",
                        help="Comma-separated model names to filter")
    parser.add_argument("--prompting-conditions", type=str, default="")
    parser.add_argument("--dataset-variants", type=str, default="")
    parser.add_argument("--patterns", type=str, default="")
    parser.add_argument("--overwrite", action="store_true")
    args = parser.parse_args()

    eval_root   = args.eval_root.resolve() / args.mode
    report_root = args.report_root.resolve() / args.mode
    csv_root = report_root / "csv"
    xlsx_root = report_root / "xlsx"

    # Determine which response-type subdirs to scan
    rt_filter = set(_csv_items(args.response_types))
    if rt_filter:
        scan_roots = [eval_root / rt for rt in sorted(rt_filter)]
    else:
        scan_roots = sorted(p for p in eval_root.iterdir() if p.is_dir()) if eval_root.exists() else []

    if not scan_roots:
        print(f"No eval directories found under: {_relpath(eval_root)}")
        return 1

    model_filter = set(_csv_items(args.models))
    op_filter    = set(_csv_items(args.prompting_conditions))
    ds_filter    = set(_csv_items(args.dataset_variants))
    rule_filter  = set(_csv_items(args.patterns))

    mode = args.mode
    eval_glob = f"eval-{mode}__*.jsonl"

    eval_files: list[Path] = []
    for scan_root in scan_roots:
        if scan_root.exists():
            eval_files.extend(scan_root.rglob(eval_glob))
    eval_files = sorted(set(eval_files))

    if not eval_files:
        print(f"No eval files found under: {_relpath(eval_root)}")
        return 1

    def _matches(p: Path) -> bool:
        import re as _re
        m_prefix = _re.match(r'^eval(?:-\w+)?__', p.stem)
        if not m_prefix:
            return False
        fields = p.stem[m_prefix.end():].split("__")
        if len(fields) < 4:
            return False
        m, op, ds, rule = fields[0], fields[1], fields[2], fields[3]
        if model_filter and m    not in model_filter: return False
        if op_filter    and op   not in op_filter:    return False
        if ds_filter    and ds   not in ds_filter:    return False
        if rule_filter  and rule not in rule_filter:  return False
        return True

    eval_files = [p for p in eval_files if _matches(p)]
    if not eval_files:
        print("No eval files matched the filters.")
        return 1

    out_path = csv_root / f"scores-{mode}.csv"
    xlsx_path = xlsx_root / f"scores-{mode}.xlsx"
    existing = [p for p in (out_path, xlsx_path) if p.exists()]
    if existing and not args.overwrite:
        print("Output already exists (use --overwrite):")
        for p in existing:
            print(f"  {_relpath(p)}")
        return 1

    records: list[dict] = []
    validation_errors: dict[str, list[str]] = {}
    n_error = 0
    for p in eval_files:
        try:
            rec = aggregate_file(p)
        except EvalValidationError as e:
            validation_errors.setdefault(e.issue, []).append(str(e))
            continue
        if rec is None:
            n_error += 1
        else:
            records.append(rec)

    if validation_errors:
        _print_eval_validation_errors(validation_errors)
        return 1

    if not records:
        print("No records aggregated.")
        return 1

    # Sort: model, prompting_condition, dataset_variant, pattern_id
    records.sort(key=lambda r: (r["model"], r["prompting_condition"], r["dataset_variant"], r["pattern_id"]))

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDNAMES)
        writer.writeheader()
        writer.writerows(records)

    print(f"Written {len(records)} rows -> {_relpath(out_path)}")
    write_excel_view(records, xlsx_path)
    if n_error:
        print(f"  ({n_error} files skipped due to errors)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
