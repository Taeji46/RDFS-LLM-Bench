"""Per-rule F1 analysis using 1-rule scenarios only.

Filters to n_rule == 1 and rule_format == "full".
Each rule's score is its mean f1_triple in single-rule scenarios,
uncontaminated by other rules' difficulty in multi-rule scenarios.

One sheet per dataset variant (rk / ls / gs / gsc / ns / nsc / rva).
Each sheet: rows = rules (rdfs2~rdfs11), columns = models.

Reads:  data/llm-eval/reports/{mode}/csv/scores-{mode}.csv
Writes:
  data/llm-eval/reports/{mode}/csv/rule_accuracy_analysis-{mode}.csv   (canonical)
  data/llm-eval/reports/{mode}/xlsx/rule_accuracy_analysis-{mode}.xlsx  (view)
"""

from __future__ import annotations

import argparse
import csv
from collections import defaultdict
from fractions import Fraction
from pathlib import Path
import sys

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

THIS_FILE = Path(__file__).resolve()
LLM_EVAL_DIR = THIS_FILE.parents[1]
PROJECT_ROOT = THIS_FILE.parents[3]
sys.path.insert(0, str(LLM_EVAL_DIR))

from shared.numeric import mean_fraction, metric_from_str, metric_to_decimal, metric_to_excel_float, metric_to_str

DEFAULT_REPORT_ROOT = PROJECT_ROOT / "data" / "llm-eval" / "reports"

ALL_RULES = ["rdfs2", "rdfs3", "rdfs5", "rdfs7", "rdfs9", "rdfs11"]
DATASET_VARIANTS = ["rk", "ls", "gs", "gsc", "ns", "nsc", "rva"]
CSV_FIELDNAMES = ["dataset_variant", "rule", "model", "f1_triple"]


def _relpath(p: Path) -> str:
    try:
        return p.relative_to(PROJECT_ROOT).as_posix()
    except ValueError:
        return str(p)


def load_scores(csv_path: Path) -> list[dict]:
    with csv_path.open(encoding="utf-8", newline="") as f:
        return list(csv.DictReader(f))


def compute(records: list[dict]) -> dict[str, dict[str, dict[str, Fraction | None]]]:
    """Returns {dataset_variant: {model: {pattern_id: mean_f1}}}.

    Only 1-rule, rule_format==full records are used.
    Each rule's score is the mean f1_triple across prompting_conditions
    (macro-average since record counts are equal across prompting_conditions).
    """
    # (dataset_variant, model, pattern_id, prompting_condition) -> [f1 values]
    fine: dict[tuple[str, str, str, str], list[Fraction]] = defaultdict(list)

    for rec in records:
        if rec.get("n_rule") != "1":
            continue
        if rec.get("rule_format") != "full":
            continue
        f1 = metric_from_str(rec.get("f1_triple", ""))
        if f1 is None:
            continue
        fine[(rec["dataset_variant"], rec["model"], rec["pattern_id"], rec["prompting_condition"])].append(f1)

    models = sorted({k[1] for k in fine})
    result: dict[str, dict[str, dict[str, Fraction | None]]] = {}
    for ds in DATASET_VARIANTS:
        result[ds] = {}
        for model in models:
            result[ds][model] = {}
            for rule in ALL_RULES:
                op_means = [
                    op_mean
                    for (d, m, r, op), v in fine.items()
                    if d == ds and m == model and r == rule and v
                    if (op_mean := mean_fraction(v)) is not None
                ]
                result[ds][model][rule] = mean_fraction(op_means) if op_means else None
    return result


def _write_sheet(wb: openpyxl.Workbook, title: str, ds_data: dict[str, dict[str, Fraction | None]], first: bool) -> None:
    if first:
        ws = wb.active
        assert ws is not None
        ws.title = title
    else:
        ws = wb.create_sheet(title=title)

    hdr_fill = PatternFill("solid", fgColor="4472C4")
    hdr_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center")
    right = Alignment(horizontal="right")

    models = sorted(ds_data.keys())

    ws.cell(row=1, column=1, value="rule").font = hdr_font
    ws.cell(row=1, column=1).fill = hdr_fill
    ws.cell(row=1, column=1).alignment = center
    for col, model in enumerate(models, 2):
        c = ws.cell(row=1, column=col, value=model)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = center

    for row, rule in enumerate(ALL_RULES, 2):
        ws.cell(row=row, column=1, value=rule).font = Font(bold=True)
        for col, model in enumerate(models, 2):
            val = ds_data[model].get(rule)
            c = ws.cell(
                row=row,
                column=col,
                value=metric_to_excel_float(val) if val is not None else "N/A",
            )
            c.alignment = right
            if val is not None:
                c.number_format = "0.000"

    ws.column_dimensions["A"].width = 10
    for col in range(2, len(models) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 26
    ws.freeze_panes = "B2"


def write_excel(data: dict[str, dict[str, dict[str, Fraction | None]]], out_path: Path) -> None:
    wb = openpyxl.Workbook()
    for i, ds in enumerate(DATASET_VARIANTS):
        _write_sheet(wb, ds, data[ds], first=(i == 0))

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Written -> {_relpath(out_path)}")


def write_csv(data: dict[str, dict[str, dict[str, Fraction | None]]], out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_FIELDNAMES)
        writer.writeheader()
        for ds in DATASET_VARIANTS:
            for rule in ALL_RULES:
                for model in sorted(data[ds]):
                    val = data[ds][model].get(rule)
                    writer.writerow({
                        "dataset_variant": ds,
                        "rule": rule,
                        "model": model,
                        "f1_triple": metric_to_str(val) if val is not None else "",
                    })
    print(f"Written CSV -> {_relpath(out_path)}")


def main() -> int:
    parser = argparse.ArgumentParser(description="Per-rule F1 analysis using 1-rule scenarios only.")
    parser.add_argument("--mode", default="strict", choices=["strict", "flex"])
    parser.add_argument("--report-root", type=Path, default=DEFAULT_REPORT_ROOT)
    parser.add_argument("--overwrite", action="store_true")
    args = parser.parse_args()

    report_root = args.report_root.resolve() / args.mode
    csv_root = report_root / "csv"
    xlsx_root = report_root / "xlsx"
    csv_path = csv_root / f"scores-{args.mode}.csv"
    csv_out = csv_root / f"rule_accuracy_analysis-{args.mode}.csv"
    xlsx_out = xlsx_root / f"rule_accuracy_analysis-{args.mode}.xlsx"

    if not csv_path.exists():
        print(f"Not found: {_relpath(csv_path)}")
        return 1
    existing = [p for p in (csv_out, xlsx_out) if p.exists()]
    if existing and not args.overwrite:
        print("Output already exists (use --overwrite):")
        for p in existing:
            print(f"  {_relpath(p)}")
        return 1

    records = load_scores(csv_path)
    data = compute(records)

    for ds in DATASET_VARIANTS:
        print(f"[{ds}]")
        for rule in ALL_RULES:
            vals = {m: data[ds][m].get(rule) for m in sorted(data[ds])}
            row_str = "  ".join(
                f"{m}={metric_to_decimal(v):.4f}" if v is not None else f"{m}=N/A"
                for m, v in vals.items()
            )
            print(f"  {rule}: {row_str}")

    write_csv(data, csv_out)
    write_excel(data, xlsx_out)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
