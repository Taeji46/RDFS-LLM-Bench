"""Scaling analysis: how f1_triple changes across 1-rule / 2-rule / 3-rule.

Averages f1_triple per (dataset_variant, model, presented_rule_type, n_rule)
across all pattern_ids, separately for each rule_format ∈ {full, def, name}.

Valid (presented_rule_type, n_rule) combinations:
  NRP: 1, 2, 3
  ARP: 1, 2, 3

Sheets (one per (rule_format, dataset_variant) combination, when data exists):
  - rule_format == "full":  rk / ls / gs / gsc / ns / nsc / rva
  - rule_format == "def":   rk-def / ls-def / gs-def / gsc-def / ns-def / nsc-def / rva-def
  - rule_format == "name":  rk-name / ls-name / gs-name / gsc-name / ns-name / nsc-name / rva-name

Cells without experimental data appear blank (e.g., gpt-oss × name on rk).

Reads:  data/llm-eval/reports/{mode}/csv/scores-{mode}.csv
Writes:
  data/llm-eval/reports/{mode}/csv/scaling_analysis-{mode}.csv   (canonical)
  data/llm-eval/reports/{mode}/xlsx/scaling_analysis-{mode}.xlsx  (view)
"""

from __future__ import annotations

import argparse
import csv
from collections import defaultdict
from fractions import Fraction
from pathlib import Path
import sys

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

THIS_FILE = Path(__file__).resolve()
LLM_EVAL_DIR = THIS_FILE.parents[1]
PROJECT_ROOT = THIS_FILE.parents[3]
sys.path.insert(0, str(LLM_EVAL_DIR))

from shared.numeric import mean_fraction, metric_from_str, metric_to_decimal, metric_to_excel_float, metric_to_str

DEFAULT_REPORT_ROOT = PROJECT_ROOT / "data" / "llm-eval" / "reports"

FAMILIES = ["NRP", "ARP"]
VALID_COMBOS = {
    "NRP": [1, 2, 3],
    "ARP": [1, 2, 3],
}
DATASET_VARIANTS = ["rk", "ls", "gs", "gsc", "ns", "nsc", "rva"]
RULE_FORMATS = ["full", "def", "name"]
CSV_FIELDNAMES = ["rule_format", "dataset_variant", "presented_rule_type", "n_rule", "model", "f1_triple"]


def _relpath(p: Path) -> str:
    try:
        return p.relative_to(PROJECT_ROOT).as_posix()
    except ValueError:
        return str(p)


def load_scores(csv_path: Path) -> list[dict]:
    with csv_path.open(encoding="utf-8", newline="") as f:
        return list(csv.DictReader(f))


def compute(
    records: list[dict], rule_format_filter: str = "full", dataset_filter: list[str] | None = None
) -> dict[str, dict[str, dict[tuple[str, int], Fraction | None]]]:
    """Returns {dataset_variant: {model: {(family, n_rule): mean_f1}}}.

    Filters records by rule_format (default "full") and optionally by dataset_variant.
    """
    datasets = dataset_filter if dataset_filter is not None else DATASET_VARIANTS
    buckets: dict[tuple[str, str, str, int], list[Fraction]] = defaultdict(list)

    for rec in records:
        if rec.get("rule_format") != rule_format_filter:
            continue
        f1 = metric_from_str(rec.get("f1_triple", ""))
        if f1 is None:
            continue
        model = rec["model"]
        ds = rec["dataset_variant"]
        if ds not in datasets:
            continue
        family = rec["presented_rule_type"]
        n_rule = int(rec["n_rule"])
        if family not in FAMILIES:
            continue
        if n_rule not in VALID_COMBOS.get(family, []):
            continue
        buckets[(ds, model, family, n_rule)].append(f1)

    models = sorted({k[1] for k in buckets})
    result: dict[str, dict[str, dict[tuple[str, int], Fraction | None]]] = {}
    for ds in datasets:
        result[ds] = {}
        for model in models:
            result[ds][model] = {}
            for family, n_rules in VALID_COMBOS.items():
                for n_rule in n_rules:
                    vals = buckets.get((ds, model, family, n_rule), [])
                    result[ds][model][(family, n_rule)] = mean_fraction(vals) if vals else None
    return result


def _write_sheet(wb: openpyxl.Workbook, title: str, ds_data: dict[str, dict[tuple[str, int], Fraction | None]], first: bool) -> None:
    if first:
        ws = wb.active
        assert ws is not None
        ws.title = title
    else:
        ws = wb.create_sheet(title=title)

    hdr_fill = PatternFill("solid", fgColor="4472C4")
    hdr_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center")
    right = Alignment(horizontal="right")

    models = sorted(ds_data.keys())

    ws.cell(row=1, column=1, value="presented_rule_type").font = hdr_font
    ws.cell(row=1, column=1).fill = hdr_fill
    ws.cell(row=1, column=1).alignment = center
    ws.cell(row=1, column=2, value="n_rule").font = hdr_font
    ws.cell(row=1, column=2).fill = hdr_fill
    ws.cell(row=1, column=2).alignment = center
    for col, model in enumerate(models, 3):
        c = ws.cell(row=1, column=col, value=model)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = center

    row = 2
    for family in FAMILIES:
        for n_rule in VALID_COMBOS[family]:
            ws.cell(row=row, column=1, value=family).font = Font(bold=True)
            ws.cell(row=row, column=2, value=n_rule).font = Font(bold=True)
            for col, model in enumerate(models, 3):
                val = ds_data[model].get((family, n_rule))
                c = ws.cell(
                    row=row,
                    column=col,
                    value=metric_to_excel_float(val) if val is not None else "",
                )
                c.alignment = right
                if val is not None:
                    c.number_format = "0.000"
            row += 1

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 10
    for col in range(3, len(models) + 3):
        ws.column_dimensions[get_column_letter(col)].width = 26
    ws.freeze_panes = "C2"


def write_excel(
    data_by_info: dict[str, dict[str, dict[str, dict[tuple[str, int], Fraction | None]]]],
    out_path: Path,
) -> None:
    """data_by_info: {rule_format: {dataset_variant: {model: {(family, n_rule): f1}}}}."""
    wb = openpyxl.Workbook()
    first = True
    for ri in RULE_FORMATS:
        ds_data_by_ds = data_by_info.get(ri, {})
        for ds in DATASET_VARIANTS:
            ds_data = ds_data_by_ds.get(ds, {})
            if not ds_data:
                continue
            # Skip sheet if all cells empty (no model has any data)
            has_data = any(
                any(v is not None for v in m.values())
                for m in ds_data.values()
            )
            if not has_data:
                continue
            title = ds if ri == "full" else f"{ds}-{ri}"
            _write_sheet(wb, title, ds_data, first=first)
            first = False

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Written -> {_relpath(out_path)}")


def write_csv(
    data_by_info: dict[str, dict[str, dict[str, dict[tuple[str, int], Fraction | None]]]],
    out_path: Path,
) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_FIELDNAMES)
        writer.writeheader()
        for rule_format in RULE_FORMATS:
            ds_data_by_ds = data_by_info.get(rule_format, {})
            for ds in DATASET_VARIANTS:
                ds_data = ds_data_by_ds.get(ds, {})
                if not ds_data:
                    continue
                for family in FAMILIES:
                    for n_rule in VALID_COMBOS[family]:
                        for model in sorted(ds_data):
                            val = ds_data[model].get((family, n_rule))
                            writer.writerow({
                                "rule_format": rule_format,
                                "dataset_variant": ds,
                                "presented_rule_type": family,
                                "n_rule": n_rule,
                                "model": model,
                                "f1_triple": metric_to_str(val) if val is not None else "",
                            })
    print(f"Written CSV -> {_relpath(out_path)}")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--mode", default="strict", choices=["strict", "flex"])
    parser.add_argument("--report-root", type=Path, default=DEFAULT_REPORT_ROOT)
    parser.add_argument("--overwrite", action="store_true")
    args = parser.parse_args()

    report_root = args.report_root.resolve() / args.mode
    csv_root = report_root / "csv"
    xlsx_root = report_root / "xlsx"
    csv_path = csv_root / f"scores-{args.mode}.csv"
    csv_out = csv_root / f"scaling_analysis-{args.mode}.csv"
    xlsx_out = xlsx_root / f"scaling_analysis-{args.mode}.xlsx"

    if not csv_path.exists():
        print(f"Not found: {_relpath(csv_path)}")
        return 1
    existing = [p for p in (csv_out, xlsx_out) if p.exists()]
    if existing and not args.overwrite:
        print("Output already exists (use --overwrite):")
        for p in existing:
            print(f"  {_relpath(p)}")
        return 1

    records = load_scores(csv_path)
    data_by_info = {
        ri: compute(records, rule_format_filter=ri) for ri in RULE_FORMATS
    }

    for ri in RULE_FORMATS:
        for ds in DATASET_VARIANTS:
            ds_data = data_by_info[ri].get(ds, {})
            if not ds_data:
                continue
            has_data = any(
                any(v is not None for v in m.values())
                for m in ds_data.values()
            )
            if not has_data:
                continue
            label = ds if ri == "full" else f"{ds}-{ri}"
            print(f"[{label}]")
            for family in FAMILIES:
                for n_rule in VALID_COMBOS[family]:
                    vals = {m: ds_data[m].get((family, n_rule)) for m in sorted(ds_data)}
                    row_str = "  ".join(
                        f"{m}={metric_to_decimal(v):.4f}" if v is not None else f"{m}=N/A"
                        for m, v in vals.items()
                    )
                    print(f"  {family} {n_rule}-rule: {row_str}")

    write_csv(data_by_info, csv_out)
    write_excel(data_by_info, xlsx_out)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
