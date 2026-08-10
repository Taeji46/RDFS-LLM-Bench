"""Compute composite metrics for LLM RDFS benchmark comparison.

Reads:
  data/llm-eval/reports/{mode}/csv/scores-{mode}.csv

Writes:
  data/llm-eval/reports/{mode}/csv/composite_metrics-{mode}.csv   (canonical)
  data/llm-eval/reports/{mode}/xlsx/composite_metrics-{mode}.xlsx  (view)

Composite metrics (each in [0, 1], higher is better):

Inference ability:
  RI  (Real-world Inference)    : f1_triple on RK, across NRP/ARP × n_rule
  SI  (Structural Inference)    : f1_triple on NS, across NRP/ARP × n_rule

Rule selection ability:
  RRS (Real-world Rule Selection): f1_rule on RK, across ARP-full × n_rule
  SRS (Structural Rule Selection): f1_rule on NS, across ARP-full × n_rule

Robustness:
  VR  (Vocabulary Robustness)   : mean min(1, GSC/NSC) across NRP/ARP × n_rule
  TR  (Token / Naming Robustness): mean min(1, GS/GSC) across NRP/ARP × n_rule

Pre-training knowledge:
  RDI (Rule Definition Independence): mean min(1, name/full) on NS across all 6 combos
    NRP:1,2,3→NRP-name/NRP-full, ARP:1,2,3→ARP-name/ARP-full
    High RDI ≈ model does not need the rule definition text (pre-training knowledge)
    Low RDI  ≈ model relies on the provided rule definition

Averaging procedure:
  For each of 6 (family × n_rule) combinations:
    (NRP:1-rule, NRP:2-rule, NRP:3-rule, ARP:1-rule, ARP:2-rule, ARP:3-rule)
  compute the mean F1 over all applicable rule_format variants and pattern_ids,
  then take the mean of these 6 values.
"""

from __future__ import annotations

import argparse
import csv
from collections import defaultdict
from fractions import Fraction
from pathlib import Path
import sys

THIS_FILE = Path(__file__).resolve()
LLM_EVAL_DIR = THIS_FILE.parents[1]
PROJECT_ROOT = THIS_FILE.parents[3]
sys.path.insert(0, str(LLM_EVAL_DIR))

from shared.numeric import mean_fraction, metric_from_str, metric_to_decimal, metric_to_excel_float, metric_to_str

DEFAULT_REPORT_ROOT = PROJECT_ROOT / "data" / "llm-eval" / "reports"

# Valid (presented_rule_type, n_rule) combinations for averaging
COMBO_KEYS = [
    ("NRP", 1),
    ("NRP", 2),
    ("NRP", 3),
    ("ARP", 1),
    ("ARP", 2),
    ("ARP", 3),
]

# For rule selection metrics, only ARP-full and ARP-name have f1_rule
RULE_EVAL_PROMPTING_CONDITIONS = {"ARP-full", "ARP-name"}


def _relpath(path: Path) -> str:
    try:
        return path.relative_to(PROJECT_ROOT).as_posix()
    except ValueError:
        return str(path)


def _load_scores(csv_path: Path) -> list[dict]:
    with csv_path.open(encoding="utf-8", newline="") as f:
        return list(csv.DictReader(f))


def compute_metrics_for_model(records: list[dict]) -> dict[str, Fraction | None]:
    """Compute all 7 composite metrics for one model's records."""

    # Index: (prompting_condition, dataset_variant, pattern_id) -> record
    idx: dict[tuple, dict] = {}
    for rec in records:
        key = (rec["prompting_condition"], rec["dataset_variant"], rec["pattern_id"])
        idx[key] = rec

    def _f1_triple(prompting_condition: str, ds: str, pattern_id: str) -> Fraction | None:
        rec = idx.get((prompting_condition, ds, pattern_id))
        return metric_from_str(rec["f1_triple"]) if rec else None

    def _f1_rule(prompting_condition: str, ds: str, pattern_id: str) -> Fraction | None:
        rec = idx.get((prompting_condition, ds, pattern_id))
        if rec is None:
            return None
        v = metric_from_str(rec.get("f1_rule", ""))
        return v if v is not None else None

    # Collect all pattern_ids per (family, n_rule)
    combo_rules: dict[tuple, set[str]] = defaultdict(set)
    for rec in records:
        family = rec["presented_rule_type"]
        n_rule = int(rec["n_rule"])
        key = (family, n_rule)
        if key in COMBO_KEYS:
            combo_rules[key].add(rec["pattern_id"])

    # Full-format prompting condition per family (RI/SI/RRS/SRS/VR/TR use full only)
    FULL_OP = {"NRP": "NRP-full", "ARP": "ARP-full"}

    def _combo_mean_f1(ds: str, metric: str = "triple") -> Fraction | None:
        """Mean over applicable (family × n_rule) combos using full-variant only.

        Returns None if any applicable combo has no data.
        """
        combo_vals: list[Fraction] = []
        for (fam, n_rule) in COMBO_KEYS:
            op = FULL_OP.get(fam)
            if op is None:
                continue
            if metric == "rule" and op not in RULE_EVAL_PROMPTING_CONDITIONS:
                continue
            rules = combo_rules.get((fam, n_rule), set())
            cell_vals: list[Fraction] = []
            for pattern_id in rules:
                if metric == "triple":
                    v = _f1_triple(op, ds, pattern_id)
                else:
                    v = _f1_rule(op, ds, pattern_id)
                if v is not None:
                    cell_vals.append(v)
            if not cell_vals:
                return None
            cell_mean = mean_fraction(cell_vals)
            if cell_mean is not None:
                combo_vals.append(cell_mean)
        return mean_fraction(combo_vals) if combo_vals else None

    # RI / SI
    ri = _combo_mean_f1("rk", "triple")
    si = _combo_mean_f1("ns", "triple")

    # RRS / SRS
    rrs = _combo_mean_f1("rk", "rule")
    srs = _combo_mean_f1("ns", "rule")

    def _robustness_mean(num_ds: str, den_ds: str) -> Fraction | None:
        """Mean min(1, num/den) over 6 (family × n_rule) combos using full-variant only.

        Returns None if any combo has no data.
        """
        combo_vals: list[Fraction] = []
        for (fam, n_rule) in COMBO_KEYS:
            op = FULL_OP.get(fam)
            if op is None:
                continue
            rules = combo_rules.get((fam, n_rule), set())
            cell_vals: list[Fraction] = []
            for pattern_id in rules:
                num = _f1_triple(op, num_ds, pattern_id)
                den = _f1_triple(op, den_ds, pattern_id)
                if num is not None and den is not None and den > 0:
                    cell_vals.append(min(Fraction(1), num / den))
            if not cell_vals:
                return None
            cell_mean = mean_fraction(cell_vals)
            if cell_mean is not None:
                combo_vals.append(cell_mean)
        return mean_fraction(combo_vals) if combo_vals else None

    # VR: min(1, GSC/NSC)
    vr = _robustness_mean("gsc", "nsc")

    # TR: min(1, GS/GSC)
    tr = _robustness_mean("gs", "gsc")

    # RDI: mean min(1, name/full) on NS across all 6 (family × n_rule) combos
    # NRP:1,2,3 → NRP-name/NRP-full, ARP:1,2,3 → ARP-name/ARP-full
    NAME_OP = {"NRP": "NRP-name", "ARP": "ARP-name"}
    rdi_combo_vals: list[Fraction] = []
    rdi_ok = True
    for (fam, n_rule) in COMBO_KEYS:
        name_op = NAME_OP.get(fam)
        full_op = FULL_OP.get(fam)
        if name_op is None or full_op is None:
            continue
        rules = combo_rules.get((fam, n_rule), set())
        cell_vals: list[Fraction] = []
        for pattern_id in rules:
            name_v = _f1_triple(name_op, "ns", pattern_id)
            full_v = _f1_triple(full_op, "ns", pattern_id)
            if name_v is not None and full_v is not None and full_v > 0:
                cell_vals.append(min(Fraction(1), name_v / full_v))
        if not cell_vals:
            rdi_ok = False
            break
        cell_mean = mean_fraction(cell_vals)
        if cell_mean is not None:
            rdi_combo_vals.append(cell_mean)
    rdi = mean_fraction(rdi_combo_vals) if (rdi_ok and rdi_combo_vals) else None

    return {
        "RI":  ri,
        "SI":  si,
        "RRS": rrs,
        "SRS": srs,
        "VR":  vr,
        "TR":  tr,
        "RDI": rdi,
    }


FIELDNAMES = ["model", "RI", "SI", "RRS", "SRS", "VR", "TR", "RDI"]


def write_excel_view(rows: list[dict], out_path: Path) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl is not installed; skipped Excel view.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "composite metrics"

    hdr_fill = PatternFill("solid", fgColor="4472C4")
    hdr_font = Font(bold=True, color="FFFFFF")
    best_fill = PatternFill("solid", fgColor="FFE699")
    center = Alignment(horizontal="center")
    right = Alignment(horizontal="right")

    for col, field in enumerate(FIELDNAMES, 1):
        cell = ws.cell(row=1, column=col, value=field)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center

    best_by_metric: dict[str, Fraction] = {}
    for field in FIELDNAMES[1:]:
        vals = [
            metric
            for row in rows
            if (metric := metric_from_str(row.get(field))) is not None
        ]
        if vals:
            best_by_metric[field] = max(vals)

    for row_idx, row in enumerate(rows, 2):
        for col, field in enumerate(FIELDNAMES, 1):
            value = row.get(field, "")
            cell = ws.cell(row=row_idx, column=col)
            if field == "model":
                cell.value = value
                cell.font = Font(bold=True)
            elif value == "":
                cell.value = ""
            else:
                metric = metric_from_str(value)
                if metric is None:
                    cell.value = ""
                    continue
                cell.value = metric_to_excel_float(metric)
                cell.number_format = "0.000"
                cell.alignment = right
                if field in best_by_metric and metric == best_by_metric[field]:
                    cell.font = Font(bold=True)
                    cell.fill = best_fill

    ws.column_dimensions["A"].width = 24
    for col in range(2, len(FIELDNAMES) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 12
    ws.freeze_panes = "B2"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Written Excel view -> {_relpath(out_path)}")


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Compute composite metrics from scores-{mode}.csv."
    )
    parser.add_argument("--mode", type=str, default="strict",
                        choices=["strict", "flex"],
                        help="Evaluation mode (default: strict)")
    parser.add_argument("--report-root", type=Path, default=DEFAULT_REPORT_ROOT)
    parser.add_argument("--overwrite", action="store_true")
    args = parser.parse_args()

    mode        = args.mode
    report_root = args.report_root.resolve() / mode
    csv_root    = report_root / "csv"
    xlsx_root   = report_root / "xlsx"
    csv_path    = csv_root / f"scores-{mode}.csv"
    out_path    = csv_root / f"composite_metrics-{mode}.csv"
    xlsx_path   = xlsx_root / f"composite_metrics-{mode}.xlsx"

    if not csv_path.exists():
        print(f"scores-{mode}.csv not found: {_relpath(csv_path)}")
        print("Run aggregate_scores.py first.")
        return 1

    existing = [p for p in (out_path, xlsx_path) if p.exists()]
    if existing and not args.overwrite:
        print("Output already exists (use --overwrite):")
        for p in existing:
            print(f"  {_relpath(p)}")
        return 1

    records = _load_scores(csv_path)
    if not records:
        print(f"scores-{mode}.csv is empty.")
        return 1

    # Group by model
    by_model: dict[str, list[dict]] = defaultdict(list)
    for rec in records:
        by_model[rec["model"]].append(rec)

    rows: list[dict] = []
    for model in sorted(by_model):
        metrics = compute_metrics_for_model(by_model[model])
        row = {"model": model}
        for k, v in metrics.items():
            row[k] = metric_to_str(v) if v is not None else ""
        rows.append(row)
        print(f"{model}:")
        for k, v in metrics.items():
            print(
                f"  {k:4s} = {metric_to_decimal(v):.6f}"
                if v is not None
                else f"  {k:4s} = N/A"
            )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDNAMES)
        writer.writeheader()
        writer.writerows(rows)

    print(f"\nWritten -> {_relpath(out_path)}")
    write_excel_view(rows, xlsx_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
