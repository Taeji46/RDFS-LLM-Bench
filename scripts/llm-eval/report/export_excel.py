"""Export per-model score views from the canonical aggregate CSV.

Reads:
  data/llm-eval/reports/{mode}/csv/scores-{mode}.csv

Writes one canonical CSV and one human-readable Excel workbook per model:
  data/llm-eval/reports/{mode}/csv/scores-{mode}__{model}.csv
  data/llm-eval/reports/{mode}/xlsx/scores-{mode}__{model}.xlsx

The per-model CSV is a filtered copy of the canonical aggregate scores.
The workbook is a browsing view with one sheet per prompting_condition.
Metric values are copied from the canonical CSV and displayed with 3 decimals.
"""

from __future__ import annotations

import argparse
import csv
from collections import defaultdict
from pathlib import Path
import sys

THIS_FILE = Path(__file__).resolve()
LLM_EVAL_DIR = THIS_FILE.parents[1]
PROJECT_ROOT = THIS_FILE.parents[3]
sys.path.insert(0, str(LLM_EVAL_DIR))

from shared.numeric import metric_to_excel_float

DEFAULT_REPORT_ROOT = PROJECT_ROOT / "data" / "llm-eval" / "reports"

PROMPTING_CONDITION_ORDER = [
    "NRP-full", "NRP-name", "NRP-def",
    "ARP-full", "ARP-name", "ARP-def",
]

DATASET_VARIANT_ORDER = ["rk", "ls", "gs", "gsc", "rva", "ns", "nsc"]


def _relpath(path: Path) -> str:
    try:
        return path.relative_to(PROJECT_ROOT).as_posix()
    except ValueError:
        return str(path)


def _rule_sort_key(pattern_id: str) -> tuple:
    """Sort by n_rule first, then constituent rule numbers numerically."""
    import re

    nums = [int(x) for x in re.findall(r"\d+", pattern_id)]
    return (len(nums), nums)


def _write_model_csv(records: list[dict], fieldnames: list[str], out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for rec in records:
            writer.writerow({k: rec.get(k, "") for k in fieldnames})


def _write_model_workbook(model: str, records: list[dict], out_path: Path) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("openpyxl is required: pip install openpyxl") from exc

    by_op: dict[str, list[dict]] = defaultdict(list)
    for rec in records:
        by_op[rec["prompting_condition"]].append(rec)

    wb = openpyxl.Workbook()
    default_ws = wb.active
    assert default_ws is not None
    wb.remove(default_ws)

    prompting_conditions = [op for op in PROMPTING_CONDITION_ORDER if op in by_op]
    for op in sorted(by_op):
        if op not in prompting_conditions:
            prompting_conditions.append(op)

    arp_ops = {"ARP-full", "ARP-name", "ARP-def"}
    sheets: list[tuple[str, str, str]] = []
    for prompting_condition in prompting_conditions:
        sheets.append((prompting_condition, prompting_condition, "f1_triple"))
        if prompting_condition in arp_ops:
            sheets.append((f"{prompting_condition}-rule", prompting_condition, "f1_rule"))

    center = Alignment(horizontal="center")
    for sheet_title, prompting_condition, metric_col in sheets:
        op_records = by_op[prompting_condition]
        pattern_ids = sorted({r["pattern_id"] for r in op_records}, key=_rule_sort_key)
        ds_variants = [
            d for d in DATASET_VARIANT_ORDER
            if any(r["dataset_variant"] == d for r in op_records)
        ]
        for d in sorted({r["dataset_variant"] for r in op_records}):
            if d not in ds_variants:
                ds_variants.append(d)

        lookup: dict[tuple[str, str], float] = {}
        for rec in op_records:
            raw = rec.get(metric_col, "")
            if raw == "" or raw is None:
                continue
            try:
                lookup[(rec["pattern_id"], rec["dataset_variant"])] = metric_to_excel_float(raw)
            except ValueError:
                continue

        is_rule_sheet = metric_col == "f1_rule"
        header_color = "70AD47" if is_rule_sheet else "4472C4"
        row_color = "E2EFDA" if is_rule_sheet else "D9E1F2"
        header_fill = PatternFill("solid", fgColor=header_color)
        header_font = Font(bold=True, color="FFFFFF")

        ws = wb.create_sheet(title=sheet_title)
        ws.freeze_panes = "B2"

        cell = ws.cell(row=1, column=1, value="pattern_id")
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9D9D9")
        cell.alignment = center

        for col_idx, ds in enumerate(ds_variants, start=2):
            cell = ws.cell(row=1, column=col_idx, value=ds)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        for row_idx, pattern_id in enumerate(pattern_ids, start=2):
            cell = ws.cell(row=row_idx, column=1, value=pattern_id)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor=row_color)

            for col_idx, ds in enumerate(ds_variants, start=2):
                cell = ws.cell(row=row_idx, column=col_idx)
                val = lookup.get((pattern_id, ds))
                if val is None:
                    cell.value = "/"
                else:
                    cell.value = val
                    cell.number_format = "0.000"
                cell.alignment = center

        ws.column_dimensions["A"].width = 16
        for col_idx in range(2, len(ds_variants) + 2):
            ws.column_dimensions[get_column_letter(col_idx)].width = 12

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def _export_model(
    model: str,
    records: list[dict],
    fieldnames: list[str],
    csv_root: Path,
    xlsx_root: Path,
    mode: str,
    overwrite: bool,
) -> bool:
    csv_path = csv_root / f"scores-{mode}__{model}.csv"
    xlsx_path = xlsx_root / f"scores-{mode}__{model}.xlsx"
    existing = [p for p in (csv_path, xlsx_path) if p.exists()]
    if existing and not overwrite:
        for p in existing:
            print(f"  SKIP (exists): {_relpath(p)}")
        return True

    _write_model_csv(records, fieldnames, csv_path)
    _write_model_workbook(model, records, xlsx_path)
    print(f"  -> {_relpath(csv_path)}")
    print(f"  -> {_relpath(xlsx_path)}")
    return True


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Export scores-{mode}.csv to per-model CSV and Excel views."
    )
    parser.add_argument("--mode", type=str, default="strict", choices=["strict", "flex"])
    parser.add_argument("--report-root", type=Path, default=DEFAULT_REPORT_ROOT)
    parser.add_argument("--models", type=str, default="",
                        help="Comma-separated model names to filter (default: all)")
    parser.add_argument("--overwrite", action="store_true")
    args = parser.parse_args()

    mode = args.mode
    report_root = args.report_root.resolve() / mode
    csv_root = report_root / "csv"
    xlsx_root = report_root / "xlsx"
    csv_path = csv_root / f"scores-{mode}.csv"

    if not csv_path.exists():
        print(f"scores-{mode}.csv not found: {_relpath(csv_path)}")
        print("Run aggregate_scores.py first.")
        return 1

    with csv_path.open(encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        fieldnames = list(reader.fieldnames or [])
        records = list(reader)

    if not records:
        print(f"scores-{mode}.csv is empty.")
        return 1

    model_filter = {s.strip() for s in args.models.split(",") if s.strip()}
    by_model: dict[str, list[dict]] = defaultdict(list)
    for rec in records:
        by_model[rec["model"]].append(rec)

    models = sorted(m for m in by_model if not model_filter or m in model_filter)
    if not models:
        print("No models matched the filter.")
        return 1

    print(f"Exporting {len(models)} model(s) ...")
    for model in models:
        _export_model(model, by_model[model], fieldnames, csv_root, xlsx_root, mode, args.overwrite)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
