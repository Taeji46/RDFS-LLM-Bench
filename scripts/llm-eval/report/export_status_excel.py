"""Export experimental status to Excel (one sheet per model).

Cell values:
  ○ = response file exists AND all UIDs (f/b/t) match the task file
  △ = response file exists, ops/rule match, but at least one UID differs
  × = task file exists but no response (experiment not yet run)
  - = no task file (combination structurally undefined, e.g. rdfs5 on gs/gsc)

Composite-metric-relevant cells are highlighted with a hatched fill:
  - ESRA-full, EMRA-full, SRA-full: all dataset columns where task exists
  - ESRA-name × 1-rule: ns column only (for RDK metric)

Reads tasks from:
  data/llm-eval/tasks/zeroshot/{op}/{dataset}/{n-rule}/task__*.json

Reads responses from:
  data/llm-eval/responses/{openai-batch,sequential}/{model}/...

Writes to:
  data/llm-eval/reports/status.xlsx
"""

from __future__ import annotations

import argparse
import re
from collections import defaultdict
from pathlib import Path

THIS_FILE = Path(__file__).resolve()
PROJECT_ROOT = THIS_FILE.parents[3]

DEFAULT_TASK_ROOT     = PROJECT_ROOT / "data" / "llm-eval" / "tasks" / "zeroshot"
DEFAULT_RESPONSE_ROOT = PROJECT_ROOT / "data" / "llm-eval" / "responses"
DEFAULT_REPORT_ROOT   = PROJECT_ROOT / "data" / "llm-eval" / "reports"

OPERATION_TYPE_ORDER = [
    "ESRA-full", "ESRA-name", "ESRA-def",
    "EMRA-full", "EMRA-name", "EMRA-def",
    "SRA-full",  "SRA-name",  "SRA-def",
]

DATASET_TYPE_ORDER = ["rk", "ls", "gs", "gsc", "rva", "ns", "nsc"]

COMPOSITE_FULL_OPS = {"ESRA-full", "EMRA-full", "SRA-full"}

_UID_RE = re.compile(r"^[a-z]-[0-9a-f]{8}$")


def _relpath(path: Path) -> str:
    try:
        return path.relative_to(PROJECT_ROOT).as_posix()
    except ValueError:
        return str(path)


def _n_rule(rule_id: str) -> int:
    return len(re.findall(r"\d+", rule_id))


def _rule_sort_key(rule_id: str) -> tuple:
    nums = [int(x) for x in re.findall(r"\d+", rule_id)]
    return (len(nums), nums)


def _is_composite_cell(op_type: str, rule_id: str, ds_type: str) -> bool:
    """Returns True if this specific cell is needed for composite metrics."""
    if op_type in COMPOSITE_FULL_OPS:
        return ds_type in {"rk", "ns", "gs", "gsc", "nsc"}
    if op_type == "ESRA-name" and _n_rule(rule_id) == 1:
        return ds_type == "ns"
    return False


def _extract_uids(fields: list[str]) -> frozenset[str]:
    """Extract all UID-shaped tokens (e.g. f-0b4f535d, b-c01c3288, t-4c7ebc76)."""
    return frozenset(f for f in fields if _UID_RE.match(f))


def _parse_task_stem(stem: str) -> dict | None:
    if not stem.startswith("task__"):
        return None
    fields = stem[len("task__"):].split("__")
    if len(fields) < 3:
        return None
    return {
        "operation_type": fields[0],
        "dataset_type":   fields[1],
        "rule_id":        fields[2],
        "uids":           _extract_uids(fields),
    }


def _parse_response_stem(stem: str) -> dict | None:
    if not stem.startswith("response__"):
        return None
    fields = stem[len("response__"):].split("__")
    if len(fields) < 4:
        return None
    return {
        "model":          fields[0],
        "operation_type": fields[1],
        "dataset_type":   fields[2],
        "rule_id":        fields[3],
        "uids":           _extract_uids(fields),
    }


# (op, rule, ds) → frozenset of UIDs from the task file
TaskIndex = dict[tuple[str, str, str], frozenset[str]]
# model → (op, rule, ds) → set of UID-frozensets (one per response file)
ResponseIndex = dict[str, dict[tuple[str, str, str], set[frozenset[str]]]]


def scan_tasks(task_root: Path) -> TaskIndex:
    """Returns {(op_type, rule_id, dataset_type): uid_frozenset}."""
    result: TaskIndex = {}
    for p in task_root.rglob("task__*.json"):
        info = _parse_task_stem(p.stem)
        if info is None:
            continue
        key = (info["operation_type"], info["rule_id"], info["dataset_type"])
        result[key] = info["uids"]
    return result


def scan_responses(response_root: Path) -> ResponseIndex:
    """Returns {model: {(op_type, rule_id, dataset_type): set[uid_frozenset]}}."""
    result: ResponseIndex = defaultdict(lambda: defaultdict(set))
    for p in response_root.rglob("response__*.jsonl"):
        info = _parse_response_stem(p.stem)
        if info is None:
            continue
        key = (info["operation_type"], info["rule_id"], info["dataset_type"])
        result[info["model"]][key].add(info["uids"])
    return result


def _cell_status(
    cell_key: tuple[str, str, str],
    task_index: TaskIndex,
    model_responses: dict[tuple, set[frozenset[str]]],
) -> str:
    """Returns '○', '△', '×', or '-'."""
    task_uids = task_index.get(cell_key)
    if task_uids is None:
        return "-"

    resp_uid_sets = model_responses.get(cell_key)
    if not resp_uid_sets:
        return "×"

    # ○ if any response file has UIDs that are a superset of the task UIDs
    # (response may omit f-uid when task has none, so use subset check)
    if any(task_uids <= resp_uids for resp_uids in resp_uid_sets):
        return "○"

    return "△"


def write_status_excel(
    task_index: TaskIndex,
    done: ResponseIndex,
    out_path: Path,
    overwrite: bool,
) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise SystemExit("openpyxl is required: pip install openpyxl")

    if out_path.exists() and not overwrite:
        print(f"Output already exists (use --overwrite): {_relpath(out_path)}")
        return

    possible = set(task_index.keys())

    # Collect all op_types, rule_ids, dataset_types
    all_op_types: set[str] = set()
    all_rule_ids_by_op: dict[str, set[str]] = defaultdict(set)
    all_ds_types: set[str] = set()

    for op, rule, ds in possible:
        all_op_types.add(op)
        all_rule_ids_by_op[op].add(rule)
        all_ds_types.add(ds)

    op_types = [op for op in OPERATION_TYPE_ORDER if op in all_op_types]
    for op in sorted(all_op_types):
        if op not in op_types:
            op_types.append(op)

    ds_types = [d for d in DATASET_TYPE_ORDER if d in all_ds_types]
    for d in sorted(all_ds_types):
        if d not in ds_types:
            ds_types.append(d)

    # --- Styles ---
    col_header_fill = PatternFill("solid", fgColor="4472C4")
    col_header_font = Font(bold=True, color="FFFFFF")

    op_group_fill = PatternFill("solid", fgColor="BFBFBF")
    op_group_font = Font(bold=True)

    composite_fill = PatternFill(patternType="gray125", fgColor="C55A11", bgColor="FFE699")
    normal_fill    = PatternFill("solid", fgColor="FFFFFF")
    na_fill        = PatternFill("solid", fgColor="F2F2F2")

    circle_font   = Font(bold=True, color="0070C0")   # ○ blue
    triangle_font = Font(bold=True, color="FF8C00")   # △ orange
    cross_font    = Font(bold=True, color="C00000")   # × red
    na_font       = Font(color="AAAAAA")              # - gray
    label_composite_font = Font(bold=True, color="843C00")
    label_normal_font    = Font(color="000000")

    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center", indent=1)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for model in sorted(done.keys()):
        model_responses = done[model]
        ws = wb.create_sheet(title=model[:31])

        # Row 1: column headers
        ws.row_dimensions[1].height = 20
        label_cell = ws.cell(row=1, column=1, value="op-type / rule-id")
        label_cell.font = Font(bold=True, color="FFFFFF")
        label_cell.fill = col_header_fill
        label_cell.alignment = center

        for col_idx, ds in enumerate(ds_types, start=2):
            cell = ws.cell(row=1, column=col_idx, value=ds)
            cell.font = col_header_font
            cell.fill = col_header_fill
            cell.alignment = center

        current_row = 2

        for op_type in op_types:
            rule_ids = sorted(all_rule_ids_by_op.get(op_type, set()), key=_rule_sort_key)
            if not rule_ids:
                continue

            # op_type group header row
            ws.row_dimensions[current_row].height = 16
            ws.merge_cells(
                start_row=current_row, start_column=1,
                end_row=current_row, end_column=len(ds_types) + 1,
            )
            gh = ws.cell(row=current_row, column=1, value=op_type)
            gh.font = op_group_font
            gh.fill = op_group_fill
            gh.alignment = left
            current_row += 1

            for rule_id in rule_ids:
                row_has_composite = any(
                    _is_composite_cell(op_type, rule_id, ds) and (op_type, rule_id, ds) in possible
                    for ds in ds_types
                )

                ws.row_dimensions[current_row].height = 15
                lc = ws.cell(row=current_row, column=1, value=rule_id)
                lc.fill      = composite_fill if row_has_composite else normal_fill
                lc.font      = label_composite_font if row_has_composite else label_normal_font
                lc.alignment = left

                for col_idx, ds in enumerate(ds_types, start=2):
                    cell_key = (op_type, rule_id, ds)
                    is_composite = _is_composite_cell(op_type, rule_id, ds) and cell_key in possible
                    status = _cell_status(cell_key, task_index, model_responses)

                    cell = ws.cell(row=current_row, column=col_idx, value=status)
                    cell.alignment = center

                    if status == "-":
                        cell.fill = na_fill
                        cell.font = na_font
                    elif status == "○":
                        cell.fill = composite_fill if is_composite else normal_fill
                        cell.font = circle_font
                    elif status == "△":
                        cell.fill = composite_fill if is_composite else normal_fill
                        cell.font = triangle_font
                    else:  # ×
                        cell.fill = composite_fill if is_composite else normal_fill
                        cell.font = cross_font

                current_row += 1

        # Column widths
        ws.column_dimensions["A"].width = 20
        for col_idx in range(2, len(ds_types) + 2):
            ws.column_dimensions[get_column_letter(col_idx)].width = 7

        ws.freeze_panes = "B2"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Written -> {_relpath(out_path)}")


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Export experimental status to Excel (one sheet per model)."
    )
    parser.add_argument("--task-root", type=Path, default=DEFAULT_TASK_ROOT,
                        help="Root of task files (default: data/llm-eval/tasks/zeroshot)")
    parser.add_argument("--response-root", type=Path, default=DEFAULT_RESPONSE_ROOT,
                        help="Root of response files (default: data/llm-eval/responses)")
    parser.add_argument("--report-root", type=Path, default=DEFAULT_REPORT_ROOT,
                        help="Output directory (default: data/llm-eval/reports)")
    parser.add_argument("--overwrite", action="store_true",
                        help="Overwrite existing output file")
    args = parser.parse_args()

    task_root     = args.task_root.resolve()
    response_root = args.response_root.resolve()
    report_root   = args.report_root.resolve()
    out_path      = report_root / "status.xlsx"

    print(f"Scanning tasks    : {_relpath(task_root)}")
    task_index = scan_tasks(task_root)
    print(f"  {len(task_index)} (op, rule, dataset) combinations defined")

    print(f"Scanning responses: {_relpath(response_root)}")
    done = scan_responses(response_root)
    print(f"  Models: {', '.join(sorted(done.keys()))}")

    write_status_excel(task_index, done, out_path, args.overwrite)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
