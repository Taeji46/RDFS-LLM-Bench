"""F1 by dataset variant — table generator.

For each (model, dataset_variant), compute mean F1 averaged over:
  - rule_format == "full" only
  - NRP × {1-rule, 2-rule, 3-rule} + ARP × {1-rule, 2-rule, 3-rule} = 6 cells
  - Each cell: mean f1_triple across all pattern_ids (implication patterns) in
               that (presented_rule_type × n_rule)
  - Final: mean of the 6 cells

Reads:  data/llm-eval/reports/{mode}/csv/scores-{mode}.csv
Writes:
  data/llm-eval/reports/{mode}/csv/f1_by_dataset-{mode}.csv   (canonical)
  data/llm-eval/reports/{mode}/xlsx/f1_by_dataset-{mode}.xlsx  (view)
"""

from __future__ import annotations

import argparse
import csv
from collections import defaultdict
from fractions import Fraction
from pathlib import Path
import sys

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

THIS_FILE = Path(__file__).resolve()
LLM_EVAL_DIR = THIS_FILE.parents[1]
PROJECT_ROOT = THIS_FILE.parents[3]
sys.path.insert(0, str(LLM_EVAL_DIR))

from shared.numeric import mean_fraction, metric_from_str, metric_to_decimal, metric_to_excel_float, metric_to_str

DEFAULT_REPORT_ROOT = PROJECT_ROOT / "data" / "llm-eval" / "reports"

DATASETS = ["rk", "ls", "rva", "gs", "gsc", "ns", "nsc"]
DATASET_LABELS = {
    "rk": "RK", "ls": "LS", "gs": "GS", "gsc": "GSC",
    "ns": "NS", "nsc": "NSC", "rva": "RVA",
}
MODEL_ORDER = [
    "gpt-oss-120b",
    "gpt-oss-20b",
    "gpt-4o-2024-08-06",
    "gpt-4o-mini-2024-07-18",
    "llama3.1-70b",
    "llama3.1-8b",
]
MODEL_LABELS = {
    "gpt-4o-2024-08-06": "GPT-4o",
    "gpt-4o-mini-2024-07-18": "GPT-4o mini",
    "gpt-oss-120b": "gpt-oss-120b",
    "gpt-oss-20b": "gpt-oss-20b",
    "llama3.1-70b": "Llama 3.1 70B",
    "llama3.1-8b": "Llama 3.1 8B",
}
FAMILIES = ["NRP", "ARP"]
N_RULES = ["1", "2", "3"]


def _relpath(p: Path) -> str:
    try:
        return p.relative_to(PROJECT_ROOT).as_posix()
    except ValueError:
        return str(p)


def load_scores(csv_path: Path) -> list[dict]:
    with csv_path.open(encoding="utf-8", newline="") as f:
        return list(csv.DictReader(f))


def compute(records: list[dict]) -> dict[str, dict[str, Fraction | None]]:
    """Returns {dataset: {model: mean_f1_over_6_cells}}."""
    # cell-level: (model, ds, family, n_rule) -> [f1 values across pattern_ids]
    cells: dict[tuple[str, str, str, str], list[Fraction]] = defaultdict(list)

    for rec in records:
        if rec.get("rule_format") != "full":
            continue
        f1 = metric_from_str(rec.get("f1_triple", ""))
        if f1 is None:
            continue
        op = rec.get("prompting_condition", "")
        if "-" not in op:
            continue
        family = op.split("-", 1)[0]  # NRP or ARP
        if family not in FAMILIES:
            continue
        n_rule = rec.get("n_rule", "")
        if n_rule not in N_RULES:
            continue
        cells[(rec["model"], rec["dataset_variant"], family, n_rule)].append(f1)

    # First average within each cell (across implication patterns)
    cell_means: dict[tuple[str, str, str, str], Fraction] = {
        k: cell_mean
        for k, v in cells.items()
        if (cell_mean := mean_fraction(v)) is not None
    }

    # Then average across the 6 cells per (model, ds)
    result: dict[str, dict[str, float | None]] = {}
    for ds in DATASETS:
        result[ds] = {}
        for model in MODEL_ORDER:
            cell_vals: list[Fraction] = []
            for fam in FAMILIES:
                for nr in N_RULES:
                    v = cell_means.get((model, ds, fam, nr))
                    if v is not None:
                        cell_vals.append(v)
            result[ds][model] = mean_fraction(cell_vals) if cell_vals else None
    return result


def write_excel(data: dict[str, dict[str, Fraction | None]], out_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "F1 by dataset"

    hdr_fill = PatternFill("solid", fgColor="4472C4")
    hdr_font = Font(bold=True, color="FFFFFF")
    best_fill = PatternFill("solid", fgColor="FFE699")
    center = Alignment(horizontal="center")
    right = Alignment(horizontal="right")

    # Header row: rows = models, columns = datasets
    ws.cell(row=1, column=1, value="LLM").font = hdr_font
    ws.cell(row=1, column=1).fill = hdr_fill
    ws.cell(row=1, column=1).alignment = center
    for col, ds in enumerate(DATASETS, 2):
        c = ws.cell(row=1, column=col, value=DATASET_LABELS[ds])
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = center

    # Determine best (per dataset / per column) for highlighting
    best_per_ds: dict[str, Fraction | None] = {}
    for ds in DATASETS:
        valid = [data[ds].get(m) for m in MODEL_ORDER if data[ds].get(m) is not None]
        best_per_ds[ds] = max(valid) if valid else None

    # Data rows
    for r, model in enumerate(MODEL_ORDER, 2):
        ws.cell(row=r, column=1, value=MODEL_LABELS[model]).font = Font(bold=True)
        ws.cell(row=r, column=1).alignment = center

        for col, ds in enumerate(DATASETS, 2):
            v = data[ds].get(model)
            cell = ws.cell(row=r, column=col)
            if v is None:
                cell.value = "--"
            else:
                cell.value = metric_to_excel_float(v)
                cell.number_format = "0.000"
                if best_per_ds[ds] is not None and v == best_per_ds[ds]:
                    cell.font = Font(bold=True)
                    cell.fill = best_fill
            cell.alignment = right

    ws.column_dimensions["A"].width = 18
    for col in range(2, len(DATASETS) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.freeze_panes = "B2"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Written -> {_relpath(out_path)}")


def write_csv(data: dict[str, dict[str, Fraction | None]], out_path: Path) -> None:
    fieldnames = ["model", *DATASETS]
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for model in MODEL_ORDER:
            row = {"model": model}
            for ds in DATASETS:
                v = data[ds].get(model)
                row[ds] = metric_to_str(v) if v is not None else ""
            writer.writerow(row)
    print(f"Written CSV -> {_relpath(out_path)}")


def main() -> int:
    parser = argparse.ArgumentParser(description="F1 by dataset variant table.")
    parser.add_argument("--mode", default="flex", choices=["strict", "flex"])
    parser.add_argument("--report-root", type=Path, default=DEFAULT_REPORT_ROOT)
    parser.add_argument("--overwrite", action="store_true")
    args = parser.parse_args()

    report_root = args.report_root.resolve() / args.mode
    csv_root = report_root / "csv"
    xlsx_root = report_root / "xlsx"
    csv_path = csv_root / f"scores-{args.mode}.csv"
    csv_out = csv_root / f"f1_by_dataset-{args.mode}.csv"
    xlsx_out = xlsx_root / f"f1_by_dataset-{args.mode}.xlsx"

    if not csv_path.exists():
        print(f"Not found: {_relpath(csv_path)}")
        return 1
    existing = [p for p in (csv_out, xlsx_out) if p.exists()]
    if existing and not args.overwrite:
        print("Output already exists (use --overwrite):")
        for p in existing:
            print(f"  {_relpath(p)}")
        return 1

    records = load_scores(csv_path)
    data = compute(records)

    # Console preview (rows = models, columns = datasets)
    print(f"=== F1 by dataset (mode={args.mode}) ===")
    header = f"{'LLM':18s}  " + "  ".join(f"{DATASET_LABELS[d]:>10s}" for d in DATASETS)
    print(header)
    for model in MODEL_ORDER:
        row = "  ".join(
            (
                f"{metric_to_decimal(data[ds][model]):10.6f}"
                if data[ds].get(model) is not None
                else f"{'--':>10s}"
            )
            for ds in DATASETS
        )
        print(f"{MODEL_LABELS[model]:18s}  {row}")

    write_csv(data, csv_out)
    write_excel(data, xlsx_out)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
