"""Estimate API budget from input-queues request files.

Reads request files from:
  data/llm-eval/requests/input-queues/{runner}/{model}_{op}_{ds}/*.jsonl

Counts:
  input  tokens — from prompt messages in each request line
  output tokens — from expected_output in the corresponding task file
                  (ARP ops also include estimated [used_rules: ...] line)

Pricing loaded from:
  scripts/llm-eval/model-pricing.json

Writes:
  data/llm-eval/reports/budget_estimate.xlsx
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

THIS_FILE = Path(__file__).resolve()
LLM_EVAL_DIR = THIS_FILE.parents[1]
PROJECT_ROOT = THIS_FILE.parents[3]
sys.path.insert(0, str(LLM_EVAL_DIR))

INPUT_QUEUE_ROOT = PROJECT_ROOT / "data" / "llm-eval" / "requests" / "input-queues"
TASK_ROOT        = PROJECT_ROOT / "data" / "llm-eval" / "tasks" / "zeroshot"
REPORT_ROOT      = PROJECT_ROOT / "data" / "llm-eval" / "reports"
PRICING_PATH     = LLM_EVAL_DIR / "model-pricing.json"

QUEUE_DIRS = ["openai-batch", "openai-compat", "ollama"]
ARP_OPS    = {"ARP-full", "ARP-name", "ARP-def"}

_RULE_LABEL_MAP = {
    "rdfs2": "ruleA", "rdfs3": "ruleB", "rdfs5": "ruleC",
    "rdfs7": "ruleD", "rdfs9": "ruleE", "rdfs11": "ruleF",
}


def _relpath(p: Path) -> str:
    try:
        return p.relative_to(PROJECT_ROOT).as_posix()
    except ValueError:
        return str(p)


# ── Token counting ────────────────────────────────────────────────

def _get_encoder():
    try:
        import tiktoken
        return tiktoken.get_encoding("o200k_base")
    except ImportError:
        return None


_ENCODER = _get_encoder()


def _count_tokens(text: str) -> int:
    if _ENCODER is not None:
        return len(_ENCODER.encode(text))
    # fallback: chars / 4
    return max(1, len(text) // 4)


# ── Task index ───────────────────────────────────────────────────

def _build_task_index() -> dict[tuple[str, str, str], list[dict]]:
    """(op, ds, rule) → list of task dicts."""
    index: dict[tuple[str, str, str], list[dict]] = {}
    for path in TASK_ROOT.rglob("task__*.json"):
        with path.open(encoding="utf-8") as f:
            data = json.load(f)
        meta  = data.get("metadata", {})
        op    = meta.get("prompting_condition", "")
        ds    = meta.get("dataset_variant", "")
        rule  = meta.get("pattern_id", "")
        tasks = data.get("tasks", [])
        if op and ds and rule and tasks:
            index[(op, ds, rule)] = tasks
    return index


def _expected_output_text(task: dict, prompting_condition: str, pattern_id: str) -> str:
    """Build the expected model output string (triples + used_rules for ARP)."""
    out = task.get("expected_output", "")
    if prompting_condition in ARP_OPS:
        # Reconstruct expected [used_rules: ...] from pattern_id
        rule_nums = re.findall(r"\d+", pattern_id)
        if prompting_condition == "ARP-def":
            labels = [_RULE_LABEL_MAP.get(f"rdfs{n}", f"rdfs{n}") for n in rule_nums]
            used = "[used_rules: " + ", ".join(labels) + "]"
        else:
            names = [f"rdfs{n}" for n in rule_nums]
            used = "[used_rules: " + ", ".join(names) + "]"
        out = out + "\n" + used if out else used
    return out


# ── Request file parsing ─────────────────────────────────────────

def _parse_request_stem(stem: str) -> dict | None:
    """Parse batch__ or seq__ stem → {model, op, ds, rule}."""
    for prefix in ("batch__", "seq__"):
        if stem.startswith(prefix):
            fields = stem[len(prefix):].split("__")
            if len(fields) >= 4:
                return {
                    "model": fields[0],
                    "op":    fields[1],
                    "ds":    fields[2],
                    "rule":  fields[3],
                }
    return None


def _input_text_from_batch_line(req: dict) -> str:
    messages = req.get("body", {}).get("messages", [])
    return "\n".join(m.get("content", "") for m in messages)


def _input_text_from_seq_line(req: dict) -> str:
    parts = []
    if req.get("system"):
        parts.append(req["system"])
    parts.append(str(req.get("input", "")))
    return "\n".join(parts)


def _input_text(req: dict, runner: str) -> str:
    if runner == "openai-batch":
        return _input_text_from_batch_line(req)
    return _input_text_from_seq_line(req)


# ── Main aggregation ─────────────────────────────────────────────

def _collect(task_index: dict) -> list[dict]:
    """Collect per-(model, op, ds, rule) token counts."""
    rows: list[dict] = []
    missing_tasks: set[tuple] = set()

    for runner in QUEUE_DIRS:
        runner_dir = INPUT_QUEUE_ROOT / runner
        if not runner_dir.exists():
            continue

        for subdir in sorted(runner_dir.iterdir()):
            if not subdir.is_dir() or subdir.name.startswith("queue"):
                continue

            for req_file in sorted(subdir.glob("*.jsonl")):
                info = _parse_request_stem(req_file.stem)
                if info is None:
                    continue
                model, op, ds, rule = info["model"], info["op"], info["ds"], info["rule"]

                tasks = task_index.get((op, ds, rule))
                if tasks is None:
                    if (op, ds, rule) not in missing_tasks:
                        print(f"  [WARN] no task file for ({op}, {ds}, {rule})")
                        missing_tasks.add((op, ds, rule))

                in_tokens  = 0
                out_tokens = 0
                n_tasks    = 0

                with req_file.open(encoding="utf-8") as f:
                    for i, line in enumerate(f):
                        line = line.strip()
                        if not line:
                            continue
                        req = json.loads(line)
                        in_tokens += _count_tokens(_input_text(req, runner))

                        if tasks and i < len(tasks):
                            out_text = _expected_output_text(tasks[i], op, rule)
                            out_tokens += _count_tokens(out_text)

                        n_tasks += 1

                rows.append({
                    "model":      model,
                    "runner":     runner,
                    "op":         op,
                    "ds":         ds,
                    "rule":       rule,
                    "n_tasks":    n_tasks,
                    "in_tokens":  in_tokens,
                    "out_tokens": out_tokens,
                })

    return rows


# ── Excel output ─────────────────────────────────────────────────

def _write_excel(rows: list[dict], pricing: dict, out_path: Path) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise SystemExit("openpyxl is required: pip install openpyxl")

    # Aggregate by model
    model_totals: dict[str, dict] = defaultdict(lambda: {"runner": "", "in_tokens": 0, "out_tokens": 0})
    for r in rows:
        m = r["model"]
        model_totals[m]["runner"]     = r["runner"]
        model_totals[m]["in_tokens"]  += r["in_tokens"]
        model_totals[m]["out_tokens"] += r["out_tokens"]

    wb = openpyxl.Workbook()

    # ── Summary sheet ──────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"

    hdr_fill = PatternFill("solid", fgColor="4472C4")
    hdr_font = Font(bold=True, color="FFFFFF")
    center   = Alignment(horizontal="center")
    right    = Alignment(horizontal="right")

    sum_headers = [
        "model", "runner",
        "input_tokens", "output_tokens",
        "input_cost_USD", "output_cost_USD", "total_cost_USD",
    ]
    for col, h in enumerate(sum_headers, 1):
        c = ws_sum.cell(row=1, column=col, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = center

    for row_i, model in enumerate(sorted(model_totals), start=2):
        t  = model_totals[model]
        p  = pricing.get(model, {})
        ip = p.get("input_per_1m",  0.0)
        op_ = p.get("output_per_1m", 0.0)
        in_cost  = t["in_tokens"]  / 1_000_000 * ip
        out_cost = t["out_tokens"] / 1_000_000 * op_

        vals = [
            model, t["runner"],
            t["in_tokens"], t["out_tokens"],
            round(in_cost, 4), round(out_cost, 4), round(in_cost + out_cost, 4),
        ]
        for col, v in enumerate(vals, 1):
            c = ws_sum.cell(row=row_i, column=col, value=v)
            if col >= 3:
                c.alignment = right
                if col >= 5:
                    c.number_format = '"$"#,##0.0000'

    ws_sum.column_dimensions["A"].width = 26
    ws_sum.column_dimensions["B"].width = 18
    for col in range(3, len(sum_headers) + 1):
        ws_sum.column_dimensions[get_column_letter(col)].width = 16

    # ── Detail sheet ───────────────────────────────────────────
    ws_det = wb.create_sheet("Detail")

    det_headers = [
        "model", "runner", "op", "ds", "rule",
        "n_tasks", "input_tokens", "output_tokens",
        "input_cost_USD", "output_cost_USD", "total_cost_USD",
    ]
    for col, h in enumerate(det_headers, 1):
        c = ws_det.cell(row=1, column=col, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = center

    for row_i, r in enumerate(sorted(rows, key=lambda x: (x["model"], x["op"], x["ds"], x["rule"])), start=2):
        p   = pricing.get(r["model"], {})
        ip  = p.get("input_per_1m",  0.0)
        op_ = p.get("output_per_1m", 0.0)
        in_cost  = r["in_tokens"]  / 1_000_000 * ip
        out_cost = r["out_tokens"] / 1_000_000 * op_

        vals = [
            r["model"], r["runner"], r["op"], r["ds"], r["rule"],
            r["n_tasks"], r["in_tokens"], r["out_tokens"],
            round(in_cost, 6), round(out_cost, 6), round(in_cost + out_cost, 6),
        ]
        for col, v in enumerate(vals, 1):
            c = ws_det.cell(row=row_i, column=col, value=v)
            if col >= 6:
                c.alignment = right
                if col >= 9:
                    c.number_format = '"$"#,##0.000000'

    for col, h in enumerate(det_headers, 1):
        ws_det.column_dimensions[get_column_letter(col)].width = max(14, len(h) + 2)

    ws_sum.freeze_panes = "A2"
    ws_det.freeze_panes = "A2"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Written -> {_relpath(out_path)}")


# ── Entry point ───────────────────────────────────────────────────

def main() -> int:
    parser = argparse.ArgumentParser(description="Estimate API budget from input-queues.")
    parser.add_argument("--out", type=Path, default=REPORT_ROOT / "budget_estimate.xlsx")
    parser.add_argument("--overwrite", action="store_true")
    args = parser.parse_args()

    out_path = args.out.resolve()
    if out_path.exists() and not args.overwrite:
        print(f"Output already exists (use --overwrite): {_relpath(out_path)}")
        return 1

    if _ENCODER is None:
        print("[WARN] tiktoken not installed — using chars/4 approximation. pip install tiktoken for accuracy.")

    print("Loading task index ...")
    task_index = _build_task_index()
    print(f"  {len(task_index)} (op, ds, rule) combinations")

    print("Loading pricing ...")
    with PRICING_PATH.open(encoding="utf-8") as f:
        pricing = {k: v for k, v in json.load(f).items() if not k.startswith("_")}

    print("Collecting token counts from input-queues ...")
    rows = _collect(task_index)
    print(f"  {len(rows)} request files processed")

    if not rows:
        print("No request files found.")
        return 1

    _write_excel(rows, pricing, out_path)

    # Print summary to stdout
    model_totals: dict[str, dict] = defaultdict(lambda: {"in": 0, "out": 0})
    for r in rows:
        model_totals[r["model"]]["in"]  += r["in_tokens"]
        model_totals[r["model"]]["out"] += r["out_tokens"]

    print("\n── Token summary ──")
    for model in sorted(model_totals):
        t  = model_totals[model]
        p  = pricing.get(model, {})
        ip = p.get("input_per_1m",  0.0)
        op_ = p.get("output_per_1m", 0.0)
        cost = t["in"] / 1e6 * ip + t["out"] / 1e6 * op_
        print(f"  {model:<30}  in={t['in']:>10,}  out={t['out']:>8,}  cost=${cost:.4f}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
