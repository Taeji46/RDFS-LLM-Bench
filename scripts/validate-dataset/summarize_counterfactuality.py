"""Aggregate per-pattern counterfactuality validation results into a summary.

Reads validation JSON files from:
  data/validation/{variant}/{n-rule}/validation__{variant}__{pattern}__...json

For each perturbed variant (LS, RVA, GS, GSC) it sums the per-pattern
entry/decision/counterfactual counts and reproduces the coverage and
counterfactuality rates reported in the paper (Table: Counterfactuality
Audit of the Perturbed Datasets).

Definitions (matching the paper's Table note):
  With N entries, D decided and C of those judged counterfactual,
    Coverage       = D / N
    Counterfactual = C / D
  The N - D undecided entries are the coverage loss.

Writes:
  data/validation/counterfactuality_summary.csv   (canonical, machine-readable)
  data/validation/counterfactuality_summary.xlsx  (human-readable view; requires openpyxl)
"""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path

THIS_FILE = Path(__file__).resolve()
PROJECT_ROOT = THIS_FILE.parents[2]
DEFAULT_VALIDATION_ROOT = PROJECT_ROOT / "data" / "validation"

# Report order follows the paper: rates increase with perturbation strength.
VARIANT_ORDER = ["ls", "rva", "gs", "gsc"]

FIELDNAMES = [
    "variant", "entries", "decided", "undecided",
    "counterfactual", "coverage", "counterfactual_rate",
]


def _relpath(path: Path) -> str:
    try:
        return path.relative_to(PROJECT_ROOT).as_posix()
    except ValueError:
        return str(path)


def summarize_variant(variant_dir: Path) -> dict | None:
    files = sorted(variant_dir.rglob("validation__*.json"))
    if not files:
        return None

    entries = decided = counterfactual = 0
    for fp in files:
        result = json.loads(fp.read_text(encoding="utf-8"))["result"]
        entries += result["total"]
        decided += result["judged"]
        counterfactual += result["counterfactual"]

    coverage = decided / entries if entries else 0.0
    cf_rate = counterfactual / decided if decided else 0.0
    return {
        "variant": variant_dir.name.upper(),
        "entries": entries,
        "decided": decided,
        "undecided": entries - decided,
        "counterfactual": counterfactual,
        "coverage": coverage,
        "counterfactual_rate": cf_rate,
        "_n_files": len(files),
    }


def write_excel_view(records: list[dict], out_path: Path) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl is not installed; skipped Excel view.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "counterfactuality"

    hdr_fill = PatternFill("solid", fgColor="4472C4")
    hdr_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center")
    right = Alignment(horizontal="right")

    for col, field in enumerate(FIELDNAMES, 1):
        cell = ws.cell(row=1, column=col, value=field)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center

    for row_idx, record in enumerate(records, 2):
        for col, field in enumerate(FIELDNAMES, 1):
            cell = ws.cell(row=row_idx, column=col, value=record[field])
            if field in {"coverage", "counterfactual_rate"}:
                cell.number_format = "0.0000"
                cell.alignment = right
            elif field in {"entries", "decided", "undecided", "counterfactual"}:
                cell.number_format = "0"
                cell.alignment = right

    widths = {"variant": 12, "entries": 12, "decided": 12, "undecided": 12,
              "counterfactual": 16, "coverage": 12, "counterfactual_rate": 20}
    for col, field in enumerate(FIELDNAMES, 1):
        ws.column_dimensions[get_column_letter(col)].width = widths.get(field, 13)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Written Excel view -> {_relpath(out_path)}")


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Summarize counterfactuality validation results per variant."
    )
    parser.add_argument("--validation-root", type=Path, default=DEFAULT_VALIDATION_ROOT)
    args = parser.parse_args()

    validation_root = args.validation_root.resolve()
    if not validation_root.exists():
        print(f"Validation root not found: {_relpath(validation_root)}")
        return 1

    records: list[dict] = []
    for variant in VARIANT_ORDER:
        variant_dir = validation_root / variant
        if not variant_dir.is_dir():
            continue
        rec = summarize_variant(variant_dir)
        if rec is not None:
            records.append(rec)

    if not records:
        print(f"No validation files found under: {_relpath(validation_root)}")
        return 1

    csv_path = validation_root / "counterfactuality_summary.csv"
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDNAMES, extrasaction="ignore")
        writer.writeheader()
        for rec in records:
            row = dict(rec)
            row["coverage"] = f"{rec['coverage']:.4f}"
            row["counterfactual_rate"] = f"{rec['counterfactual_rate']:.4f}"
            writer.writerow(row)
    print(f"Written {len(records)} rows -> {_relpath(csv_path)}")

    xlsx_path = validation_root / "counterfactuality_summary.xlsx"
    write_excel_view(records, xlsx_path)

    # Console echo for quick inspection.
    print()
    print(f"{'variant':8s}{'entries':>9}{'decided':>9}{'counterfact':>13}"
          f"{'coverage':>11}{'cf_rate':>10}")
    for rec in records:
        print(f"{rec['variant']:8s}{rec['entries']:>9}{rec['decided']:>9}"
              f"{rec['counterfactual']:>13}{rec['coverage']*100:>10.2f}%"
              f"{rec['counterfactual_rate']*100:>9.2f}%")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
